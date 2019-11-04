import openpyxl
import json
import os

wb = openpyxl.load_workbook(filename="xlsx2json/data.xlsx", read_only=True)
sheet = wb['data_sheet01']

# ヘッダ部取得
columns = []
for index in range(1, sheet.max_column+1):
    columns.append(sheet.cell(row=1, column=index).value)

# データ部取得
dict = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    row_data = {}
    for cell in row:
        row_data[columns[cell.column-1]] = cell.value
    dict.append(row_data)

os.makedirs("out")
json_file = open('out/data.json', 'w')
json.dump(dict, json_file, ensure_ascii=False, indent=2)